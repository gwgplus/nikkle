#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
匯出功能模組
根據 frmExport.cs 和 frmExport.Designer.cs 實作
"""

import sys
import os
import csv
import json
from datetime import datetime, timedelta
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QGridLayout, QPushButton, QLabel, 
                             QDateTimeEdit, QGroupBox, QRadioButton, QButtonGroup,
                             QFileDialog, QMessageBox, QFrame)
from PyQt5.QtCore import Qt, QDate, QTime, QDateTime
from PyQt5.QtGui import QFont
from database_manager import DatabaseManager
from models import Ocrlog


class ExportWindow(QMainWindow):
    """匯出視窗"""
    
    def __init__(self,default_dir:str ,db_manager:DatabaseManager=None):
        super().__init__()
       
        self.default_dir = default_dir
        # 嘗試初始化資料庫管理器
        try:

            self.db_manager = db_manager    
        except Exception as e:
            print(f"資料庫初始化失敗: {e}")
            self.db_manager = None
        
        self.init_ui()
        
    def init_ui(self):
        """初始化使用者介面"""
        self.setWindowTitle("匯出")
        self.setFixedSize(565, 396)
        
        # 中央 widget
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # 主佈局
        main_layout = QVBoxLayout(central_widget)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # 匯出選項群組
        self.create_export_options_group(main_layout)
        
        # 按鈕區域
        self.create_buttons_area(main_layout)
        
        # 日期選擇區域
        self.create_date_selection_area(main_layout)
        
    def create_export_options_group(self, parent_layout):
        """創建匯出選項群組"""
        group_box = QGroupBox("匯出選項")
        group_box.setFont(QFont("新細明體", 12))
        
        layout = QHBoxLayout(group_box)
        
        # 匯出格式群組
        format_group = QGroupBox("匯出格式")
        format_group.setFont(QFont("新細明體", 10))
        format_layout = QVBoxLayout(format_group)
        
        self.format_button_group = QButtonGroup()
        self.rd_xlsx = QRadioButton("XLSX")
        self.rd_xlsx.setChecked(True)
        self.rd_csv = QRadioButton("CSV")
        
        self.format_button_group.addButton(self.rd_xlsx, 0)
        self.format_button_group.addButton(self.rd_csv, 1)
        
        format_layout.addWidget(self.rd_xlsx)
        format_layout.addWidget(self.rd_csv)
        
        # 分隔字元群組
        separator_group = QGroupBox("分隔字元")
        separator_group.setFont(QFont("新細明體", 10))
        separator_layout = QVBoxLayout(separator_group)
        
        self.separator_button_group = QButtonGroup()
        self.rd_comma = QRadioButton("逗號")
        self.rd_comma.setChecked(True)
        self.rd_tab = QRadioButton("Tab")
        
        self.separator_button_group.addButton(self.rd_comma, 0)
        self.separator_button_group.addButton(self.rd_tab, 1)
        
        separator_layout.addWidget(self.rd_comma)
        separator_layout.addWidget(self.rd_tab)
        
        layout.addWidget(format_group)
        layout.addWidget(separator_group)
        
        parent_layout.addWidget(group_box)
        
    def create_buttons_area(self, parent_layout):
        """創建按鈕區域"""
        buttons_layout = QHBoxLayout()
        
        # 匯出今天報表按鈕
        self.btn_export_today = QPushButton("匯出今天報表")
        self.btn_export_today.setFont(QFont("新細明體", 14))
        self.btn_export_today.setFixedSize(250, 87)
        self.btn_export_today.clicked.connect(self.export_today)
        
        # 依條件匯出按鈕
        self.btn_export_range = QPushButton("依條件匯出")
        self.btn_export_range.setFont(QFont("新細明體", 14))
        self.btn_export_range.setFixedSize(250, 87)
        self.btn_export_range.clicked.connect(self.export_range)
        
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.btn_export_today)
        buttons_layout.addStretch()
        buttons_layout.addWidget(self.btn_export_range)
        buttons_layout.addStretch()
        
        parent_layout.addLayout(buttons_layout)
        
    def create_date_selection_area(self, parent_layout):
        """創建日期選擇區域"""
        date_layout = QHBoxLayout()
        
        # 從日期
        label_from = QLabel("從")
        label_from.setFont(QFont("新細明體", 12))
        
        self.date_from = QDateTimeEdit()
        self.date_from.setDate(QDate.currentDate())
        self.date_from.setDisplayFormat("yyyy/MM/dd")
        self.date_from.setFixedSize(200, 27)
        
        # 到日期
        label_to = QLabel("到")
        label_to.setFont(QFont("新細明體", 12))
        
        self.date_to = QDateTimeEdit()
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setDisplayFormat("yyyy/MM/dd")
        self.date_to.setFixedSize(200, 27)
        
        date_layout.addWidget(label_from)
        date_layout.addWidget(self.date_from)
        date_layout.addWidget(label_to)
        date_layout.addWidget(self.date_to)
        date_layout.addStretch()
        
        parent_layout.addLayout(date_layout)
        
    def export_today(self):
        """匯出今天報表"""
        try:
            # 設定檔案名稱
            today = datetime.now().strftime("%Y-%m-%d")
            file_extension = ".xlsx" if self.rd_xlsx.isChecked() else ".csv"
            default_filename = f"{today}{file_extension}"
            
            # 選擇檔案
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "儲存檔案",
                os.path.join(self.default_dir, default_filename),
                f"*.{file_extension[1:]}|*.{file_extension[1:]}"
            )
            
            if not file_path:
                return
                
            # 查詢今天的資料
            today_start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
            today_end = today_start + timedelta(days=1)
            
            if self.db_manager is None:
                QMessageBox.warning(self, "警告", "資料庫未初始化，無法匯出資料")
                return
                
            logs = self.db_manager.get_ocr_logs_by_date_range(today_start, today_end)
            
            if not logs:
                QMessageBox.information(self, "提示", "今天尚未有資料可供輸出")
                return
                
            # 匯出檔案
            if self.rd_xlsx.isChecked():
                self.export_excel(file_path, logs)
            else:
                self.export_csv(file_path, logs)
                
            QMessageBox.information(self, "完成", "輸出完成!!")
            
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"發生錯誤: {str(e)}")
            
    def export_range(self):
        """依條件匯出"""
        try:
            # 設定檔案名稱
            from_date = self.date_from.date().toPyDate()
            to_date = self.date_to.date().toPyDate()
            file_extension = ".xlsx" if self.rd_xlsx.isChecked() else ".csv"
            default_filename = f"{from_date.strftime('%Y%m%d')}-{to_date.strftime('%y%m%d')}{file_extension}"
            
            # 選擇檔案
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "儲存檔案",
                os.path.join(self.default_dir, default_filename),
                f"*.{file_extension[1:]}|*.{file_extension[1:]}"
            )
            
            if not file_path:
                return
                
            # 查詢指定日期範圍的資料
            start_time = datetime.combine(from_date, datetime.min.time())
            end_time = datetime.combine(to_date, datetime.min.time()) + timedelta(days=1)
            
            if self.db_manager is None:
                QMessageBox.warning(self, "警告", "資料庫未初始化，無法匯出資料")
                return
                
            logs = self.db_manager.get_ocr_logs_by_date_range(start_time, end_time)
            
            if not logs:
                QMessageBox.information(self, "提示", "您所選擇的區間尚未有資料可供輸出")
                return
                
            # 匯出檔案
            if self.rd_xlsx.isChecked():
                self.export_excel(file_path, logs)
            else:
                self.export_csv(file_path, logs)
                
            QMessageBox.information(self, "完成", "輸出完成!!")
            
        except Exception as e:
            QMessageBox.critical(self, "錯誤", f"發生錯誤: {str(e)}")
            
    def export_csv(self, file_path, logs):
        """匯出 CSV 檔案"""
        try:
            # 設定分隔字元
            separator = "," if self.rd_comma.isChecked() else "\t"
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                # 寫入標題行
                headers = [
                    "日期", "時間", "第一次照合", "第二次照合", "電腦識別", 
                    "電腦判定結果", "人工判別", "Err 原因", "操作者1", "操作者2", 
                    "外觀判定", "NG原因", "磯原品", "重工件"
                ]
                
                if self.rd_tab.isChecked():
                    csvfile.write("\t".join(headers) + "\n")
                else:
                    csvfile.write(",".join(headers) + "\n")
                
                # 寫入資料行
                for log in logs:
                    row_data = self.process_log_data(log)
                    csvfile.write(separator.join(row_data) + "\n")
                    
        except Exception as e:
            raise Exception(f"CSV 匯出失敗: {str(e)}")
            
    def export_excel(self, file_path, logs):
        """匯出 Excel 檔案"""
        try:
            # 嘗試匯入 openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                QMessageBox.warning(
                    self, 
                    "警告", 
                    "未安裝 openpyxl 套件，無法匯出 Excel 檔案。\n請執行: pip install openpyxl"
                )
                return
                
            wb = Workbook()
            ws = wb.active
            ws.title = "export"
            
            # 設定標題行
            headers = [
                "日期", "時間", "第一次照合", "第二次照合", "電腦識別", 
                "電腦判定結果", "人工判別", "Err 原因", "操作者1", "操作者2", 
                "外觀判定", "NG原因", "磯原品", "重工件"
            ]
            
            # 寫入標題並設定樣式
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(color="FFFFFF", bold=True)
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20
            
            # 寫入資料
            for row_idx, log in enumerate(logs, 2):
                row_data = self.process_log_data(log)
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # 設定日期和時間格式
                    if col_idx == 1:  # 日期
                        cell.number_format = "yyyy/mm/dd"
                    elif col_idx == 2:  # 時間
                        cell.number_format = "hh:mm:ss"
                    
                    # 設定錯誤資料的顏色
                    if log.get('source') != log.get('ocr_result'):
                        if col_idx in [4, 5]:  # 第二次照合、電腦識別
                            cell.font = Font(color="FF0000")
                    
                    if log.get('judgment', 0) != 0:
                        if col_idx in [6, 7, 8]:  # 電腦判定結果、人工判別、Err 原因
                            cell.font = Font(color="FF0000")
            
            wb.save(file_path)
            
        except Exception as e:
            raise Exception(f"Excel 匯出失敗: {str(e)}")
            
    def process_log_data(self, log):
        """處理日誌資料，轉換為匯出格式"""
        # 判斷錯誤類型和結果
        str_err = ""
        str_err_select = ""
        str_result = "OK"
        
        judgment = log.get('Judgment', 0)
        if judgment == 1:
            str_err = "退回"
            str_result = "NG"
        elif judgment == 2:
            str_err = "允收"
            str_err_select = "無法辨識"
            str_result = "ERR"
        elif judgment == 3:
            str_err = "允收"
            str_err_select = "辨識錯誤"
            str_result = "ERR"
        elif judgment == 4:
            str_err = "允收"
            str_err_select = "摺痕"
            str_result = "ERR"
        
        # 外觀判定
        is_exterior_ok = log.get('IsExteriorOK')
        exterior_result = ""
        if is_exterior_ok is not None:
            exterior_result = "OK" if is_exterior_ok else "NG"
        
        # NG 原因
        str_reasons = ["", "氧化", "漏氣", "異物", "孔洞異常"]
        err_reason = log.get('ExteriorErrReason', 0) or 0
        ng_reason = "" if is_exterior_ok else str_reasons[err_reason]
        
        # 類別判定
        exterior_class = log.get('ExteriorClass', 0) or 0
        is_iso = "V" if (exterior_class & 0x01) == 0x01 else ""
        is_heavy = "V" if (exterior_class & 0x02) == 0x02 else ""
        
        # 操作者處理
        processor_str = log.get('Processor', '') or ""
        account = log.get('Account', '') or ""
        
        operator1 = ""
        operator2 = ""
        
        if processor_str:
            processors = [p.strip() for p in processor_str.split(',') if p.strip()]
            
            # 找到當前帳號
            if account and account in processors:
                operator1 = account
                processors.remove(account)
            
            # 設定第二個操作者
            if processors:
                operator2 = processors[0]
        
        # 處理時間格式
        time_str = log.get('Time', '')
        if time_str:
            try:
                # 如果是字串格式，先解析
                if isinstance(time_str, str):
                    from datetime import datetime
                    time_obj = datetime.fromisoformat(time_str.replace('Z', '+00:00'))
                else:
                    time_obj = time_str
                
                date_str = time_obj.strftime('%Y/%m/%d')
                time_only = time_obj.strftime('%H:%M:%S')
            except:
                date_str = ""
                time_only = ""
        else:
            date_str = ""
            time_only = ""
        
        # 組裝資料陣列
        row_data = [
            date_str,                   # 日期
            time_only,                  # 時間
            log.get('Source', ''),      # 第一次照合
            log.get('Source', ''),      # 第二次照合
            log.get('OCRResult', ''),   # 電腦識別
            str_result,                 # 電腦判定結果
            str_err,                    # 人工判別
            str_err_select,             # Err 原因
            operator1,                  # 操作者1
            operator2,                  # 操作者2
            exterior_result,            # 外觀判定
            ng_reason,                  # NG原因
            is_iso,                     # 磯原品
            is_heavy                    # 重工件
        ]
        
        return row_data


def main():
    """主程式"""
    app = QApplication(sys.argv)
    
    # 設定應用程式字體
    font = QFont("新細明體", 12)
    app.setFont(font)
    
    window = ExportWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
